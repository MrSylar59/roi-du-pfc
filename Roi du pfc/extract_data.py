from openpyxl import load_workbook

wb = load_workbook(filename="data_raw.xlsx", read_only=True)
ws = wb["query_result"]

def createFile(name, col):
    file = open(name, "w")
    nextInfo = 1000
    tot = 0

    for i in range(2, ws.max_row):
        val = ws.cell(row=i, column=col).value
        nextInfo -= 1
        tot += 1

        if val != 0:
            if val == 1:
                file.write("Pierre\n")
            elif val == 2:
                file.write("Feuille\n")
            else:
                file.write("Ciseaux\n")
        if nextInfo == 0:
            nextInfo = 1000
            print("Processed "+str(tot)+" out of "+str(ws.max_row)+" lines ("+str(tot/ws.max_row * 100)+"%)")
        
    file.close()

createFile("trainData.txt", 3)
createFile("testData.txt", 4)
